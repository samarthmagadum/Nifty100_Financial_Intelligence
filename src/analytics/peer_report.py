import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.database.database_utils import get_connection


def load_peer_percentiles():

    print()
    print("=" * 60)
    print("LOADING PEER PERCENTILES")
    print("=" * 60)

    conn = get_connection()

    query = """
    SELECT *
    FROM peer_percentiles
    """

    df = pd.read_sql(query, conn)

    conn.close()

    print()
    print("Rows :", len(df))
    print()
    print(df.head())

    return df


def load_peer_groups():

    print()
    print("=" * 60)
    print("LOADING PEER GROUPS")
    print("=" * 60)

    conn = get_connection()

    query = """
    SELECT
        peer_group_name,
        company_id,
        is_benchmark
    FROM peer_groups
    """

    peer_groups = pd.read_sql(query, conn)

    conn.close()

    print()
    print(peer_groups.head())

    return peer_groups


def load_company_names():

    print()
    print("=" * 60)
    print("LOADING COMPANY NAMES")
    print("=" * 60)

    conn = get_connection()

    query = """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """

    companies = pd.read_sql(query, conn)

    conn.close()

    print()
    print(companies.head())

    return companies


def prepare_value_table(df):

    print()
    print("=" * 60)
    print("PREPARING VALUE TABLE")
    print("=" * 60)

    value_table = df.pivot_table(

        index=[
            "company_id",
            "year",
            "broad_sector"
        ],

        columns="metric",

        values="value"

    ).reset_index()

    value_table.columns.name = None

    print()
    print("Rows :", len(value_table))
    print()

    print(value_table.head())

    return value_table

def prepare_percentile_table(df):

    print()
    print("=" * 60)
    print("PREPARING PERCENTILE TABLE")
    print("=" * 60)

    percentile_table = df.pivot_table(

        index=[
            "company_id",
            "year",
            "broad_sector"
        ],

        columns="metric",

        values="percentile_rank"

    ).reset_index()

    percentile_table.columns.name = None

    # Rename metric columns
    new_columns = []

    for col in percentile_table.columns:

        if col in [
            "company_id",
            "year",
            "broad_sector"
        ]:

            new_columns.append(col)

        else:

            new_columns.append(f"{col}_percentile")

    percentile_table.columns = new_columns

    print()
    print("Rows :", len(percentile_table))
    print()

    print(percentile_table.head())

    print()
    print(percentile_table.columns)

    return percentile_table


def merge_value_percentiles(
    value_table,
    percentile_table
):

    print()
    print("=" * 60)
    print("MERGING VALUE + PERCENTILE TABLES")
    print("=" * 60)

    merged_table = value_table.merge(

        percentile_table,

        on=[
            "company_id",
            "year",
            "broad_sector"
        ],

        how="inner"

    )

    print()
    print("Rows :", len(merged_table))
    print()

    print(merged_table.head())

    print()
    print(merged_table.columns)

    return merged_table


def prepare_peer_table(df):

    print()
    print("=" * 60)
    print("PREPARING PEER TABLE")
    print("=" * 60)

    peer_table = df.pivot_table(

        index=[
            "company_id",
            "year",
            "broad_sector"
        ],

        columns="metric",

        values="percentile_rank"

    ).reset_index()

    peer_table.columns.name = None

    print()
    print("Rows :", len(peer_table))
    print()

    print(peer_table.head())

    return peer_table


def merge_peer_information(

    merged_table,
    peer_groups,
    companies

):

    print()
    print("=" * 60)
    print("MERGING PEER INFORMATION")
    print("=" * 60)

    peer_table = merged_table.merge(

        peer_groups,

        on="company_id",

        how="inner"

    )

    peer_table = peer_table.merge(

        companies,

        on="company_id",

        how="left"

    )

    columns = [

        "peer_group_name",

        "company_id",

        "company_name",

        "year",

        "is_benchmark"

    ]

    other_cols = [

        col

        for col in peer_table.columns

        if col not in columns

    ]

    peer_table = peer_table[
        columns + other_cols
    ]

    print()

    print("Rows :", len(peer_table))

    print()

    print(peer_table.head())

    print()

    print("Columns")

    print(peer_table.columns)

    return peer_table


def create_peer_report(peer_table):

    print()
    print("=" * 60)
    print("CREATING PEER COMPARISON EXCEL")
    print("=" * 60)

    output_file = "output/peer_comparison.xlsx"

    print()
    print("Columns in Peer Table")
    print(peer_table.columns)

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        peer_groups = sorted(
            peer_table["peer_group_name"].dropna().unique()
        )

        for group in peer_groups:

            print("Writing Sheet :", group)

            group_df = peer_table[
                peer_table["peer_group_name"] == group
            ].copy()

            numeric_columns = group_df.select_dtypes(
                include="number"
            ).columns

            median_row = {}

            median_row["peer_group_name"] = group
            median_row["company_id"] = "Median"
            median_row["company_name"] = ""
            median_row["year"] = ""
            median_row["is_benchmark"] = ""
            median_row["broad_sector"] = ""

            for col in numeric_columns:

                median_row[col] = round(
                    group_df[col].median(),
                    2
                )

            group_df = pd.concat(

                [
                    group_df,
                    pd.DataFrame([median_row])
                ],

                ignore_index=True

            )

            column_order = [

                "peer_group_name",
                "company_id",
                "company_name",
                "year",
                "broad_sector",
                "is_benchmark"

            ]

            remaining = [

                col

                for col in group_df.columns

                if col not in column_order

            ]

            group_df = group_df[
                column_order + remaining
            ]

            group_df.to_excel(

                writer,

                sheet_name=group[:31],

                index=False

            )

    print()
    print("=" * 60)
    print("WORKBOOK CREATED")
    print("=" * 60)

    print(output_file)


def apply_formatting(output_file):

    print()
    print("=" * 60)
    print("APPLYING FORMATTING")
    print("=" * 60)

    wb = load_workbook(output_file)

    # Colours
    green = PatternFill(fill_type="solid", start_color="90EE90")
    yellow = PatternFill(fill_type="solid", start_color="FFF59D")
    red = PatternFill(fill_type="solid", start_color="FF9999")
    gold = PatternFill(fill_type="solid", start_color="FFD966")

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    percentile_columns = [

        "return_on_equity_pct_percentile",
        "net_profit_margin_pct_percentile",
        "debt_to_equity_percentile",
        "free_cash_flow_cr_percentile",
        "interest_coverage_percentile",
        "asset_turnover_percentile",
        "revenue_cagr_5yr_percentile",
        "pat_cagr_5yr_percentile",
        "composite_quality_score_percentile"

    ]

    for ws in wb.worksheets:

        # Freeze first row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]

        # Format header
        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        percentile_cols = []

        for idx, column in enumerate(headers, start=1):

            if column in percentile_columns:
                percentile_cols.append(idx)

        benchmark_col = headers.index("is_benchmark") + 1

        # Hide benchmark column
        ws.column_dimensions[
            get_column_letter(benchmark_col)
        ].hidden = True

        for row in range(2, ws.max_row + 1):

            benchmark = ws.cell(
                row=row,
                column=benchmark_col
            ).value

            # Center all data
            for cell in ws[row]:
                cell.alignment = center

            # Highlight benchmark row
            if benchmark == 1:

                for cell in ws[row]:
                    cell.fill = gold

                # Skip percentile colouring for benchmark row
                continue

            # Apply percentile colours
            for col in percentile_cols:

                cell = ws.cell(row=row, column=col)

                if isinstance(cell.value, (int, float)):

                    if cell.value >= 75:

                        cell.fill = green

                    elif cell.value >= 25:

                        cell.fill = yellow

                    else:

                        cell.fill = red

        # Auto-fit columns
        for column_cells in ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = max_length + 3

    wb.save(output_file)

    print()
    print("=" * 60)
    print("FORMATTING APPLIED SUCCESSFULLY")
    print("=" * 60)
    print()
    print("File :", output_file)


if __name__ == "__main__":

    print()
    print("=" * 60)
    print("PEER COMPARISON REPORT")
    print("=" * 60)

    # Load data
    percentiles = load_peer_percentiles()

    peer_groups = load_peer_groups()

    value_table = prepare_value_table(percentiles)

    percentile_table = prepare_percentile_table(percentiles)

    print(value_table.columns)

    print()
    print("Peer Groups Shape :", peer_groups.shape)
    print("Unique Companies :", peer_groups["company_id"].nunique())
    print("Peer Groups :")
    print(peer_groups["peer_group_name"].unique())

    companies = load_company_names()

    merged_table = merge_value_percentiles(
    value_table,
    percentile_table
    )

    peer_table = merge_peer_information(
    merged_table,
    peer_groups,
    companies
    )

    create_peer_report(peer_table)

    apply_formatting("output/peer_comparison.xlsx")