import pandas as pd
import yaml

from datetime import datetime
from src.database.database_utils import get_connection

from src.screener.score import winsorize
from src.screener.score import normalize
from src.screener.score import normalize_inverse

from src.screener.excel_export import export_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

GREEN_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

RED_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)


def load_config():

    print("=" * 60)
    print("LOADING CONFIGURATION")
    print("=" * 60)

    with open(
        "config/screener_config.yaml",
        "r"
    ) as file:

        config = yaml.safe_load(file)

    return config


def load_financial_ratios():

    print()
    print("=" * 60)
    print("LOADING FINANCIAL RATIOS")
    print("=" * 60)

    conn = get_connection()

    ratios_query = """
    SELECT
        f.*,
        m.pe_ratio,
        m.pb_ratio,
        m.dividend_yield_pct
    FROM financial_ratios f

    LEFT JOIN market_cap m
    ON f.company_id = m.company_id
    AND CAST(
        REPLACE(
            REPLACE(f.year,'Mar ',''),
            'Dec ',''
        ) AS INTEGER
    ) = m.year
   """

    df = pd.read_sql(ratios_query, conn)

    sectors_query = """
    SELECT
        company_id,
        broad_sector
    FROM sectors
    """

    sectors = pd.read_sql(sectors_query, conn)

    conn.close()

    df = df.merge(
        sectors,
        on="company_id",
        how="left"
    )

    print()

    print("Rows :", len(df))
    print("Sector Records :", df["broad_sector"].notna().sum())

    print()
    print("=" * 60)
    print("AVAILABLE COLUMNS")
    print("=" * 60)
    print(df.columns.tolist())

    return df


def winsorize_score(series):

    lower = series.quantile(0.10)
    upper = series.quantile(0.90)

    clipped = series.clip(lower, upper)

    if upper == lower:
        return pd.Series(50, index=series.index)

    score = (
        (clipped - lower)
        /
        (upper - lower)
    ) * 100

    return score.fillna(0)



def calculate_composite_score(df):

    print()
    print("=" * 60)
    print("CALCULATING COMPOSITE SCORE")
    print("=" * 60)

    # ---------- Profitability ----------
    roe_score = winsorize_score(df["return_on_equity_pct"])

    npm_score = winsorize_score(df["net_profit_margin_pct"])

    # ---------- Cash Quality ----------
    fcf_score = winsorize_score(df["free_cash_flow_cr"])

    # Positive Free Cash Flow Flag
    fcf_flag = (df["free_cash_flow_cr"] > 0).astype(int) * 100

    # ---------- Growth ----------
    revenue_score = winsorize_score(df["revenue_cagr_5yr"])  

    pat_score = winsorize_score(df["pat_cagr_5yr"])

    # ---------- Leverage ----------
    debt_score = winsorize_score(
    1 / (1 + df["debt_to_equity"].fillna(0))
    )

    icr_score = winsorize_score(
    df["interest_coverage"].fillna(0)
       )

    # Weighted Composite Score
    df["composite_quality_score"] = (

        roe_score * 0.20 +

        npm_score * 0.15 +

        fcf_score * 0.20 +

        fcf_flag * 0.10 +

        revenue_score * 0.15 +

        pat_score * 0.10 +

        debt_score * 0.05 +

        icr_score * 0.05

    )

    df["composite_quality_score"] = (
    df.groupby("broad_sector")["composite_quality_score"]
      .transform(winsorize_score)
    )

    return df


def apply_filters(df, config):

    filters = config["filters"]

    print()
    print("Current Filters")

    for key, value in filters.items():
        print(key, ":", value)

    print()
    print("=" * 60)
    print("APPLYING FILTERS")
    print("=" * 60)

    # -----------------------------
    # ROE
    # -----------------------------
    if "roe_min" in filters:

        df = df[
            df["return_on_equity_pct"] >= filters["roe_min"]
        ]

        print("After ROE Filter :", len(df))

    # -----------------------------
    # Debt to Equity
    # -----------------------------
    if "debt_to_equity_max" in filters:

        debt_limit = filters["debt_to_equity_max"]

        df = df[
            (df["broad_sector"] == "Financials") |
            (df["debt_to_equity"] <= debt_limit)
        ]

        print("After Debt Filter :", len(df))

    # -----------------------------
    # Free Cash Flow
    # -----------------------------
    if "free_cash_flow_min" in filters:

        df = df[
            df["free_cash_flow_cr"] >= filters["free_cash_flow_min"]
        ]

        print("After FCF Filter :", len(df))

    # -----------------------------
    # Revenue CAGR
    # -----------------------------
    if "revenue_cagr_5yr_min" in filters:

        df = df[
            df["revenue_cagr_5yr"] >=
            filters["revenue_cagr_5yr_min"]
        ]

        print("After Revenue CAGR :", len(df))

    # -----------------------------
    # PAT CAGR
    # -----------------------------
    if "pat_cagr_5yr_min" in filters:

        df = df[
            df["pat_cagr_5yr"] >=
            filters["pat_cagr_5yr_min"]
        ]

        print("After PAT CAGR :", len(df))

    # -----------------------------
    # OPM
    # -----------------------------
    if "operating_profit_margin_min" in filters:

        df = df[
            df["operating_profit_margin_pct"] >=
            filters["operating_profit_margin_min"]
        ]

        print("After OPM Filter :", len(df))

    # -----------------------------
    # PE Ratio
    # -----------------------------
    if "pe_ratio_max" in filters:

        df = df[
            df["pe_ratio"] <= filters["pe_ratio_max"]
        ]

        print("After PE Filter :", len(df))

    # -----------------------------
    # PB Ratio
    # -----------------------------
    if "pb_ratio_max" in filters:

        df = df[
            df["pb_ratio"] <= filters["pb_ratio_max"]
        ]

        print("After PB Filter :", len(df))

    # -----------------------------
    # Dividend Yield
    # -----------------------------
    if "dividend_yield_min" in filters:

        df = df[
            df["dividend_yield_pct"] >=
            filters["dividend_yield_min"]
        ]

        print("After Dividend Yield Filter :", len(df))

    # -----------------------------
    # Dividend Payout
    # -----------------------------
    if "dividend_payout_max" in filters:

        df = df[
            df["dividend_payout_ratio_pct"] <=
            filters["dividend_payout_max"]
        ]

        print("After Dividend Payout Filter :", len(df))

    df = calculate_composite_score(df)

    df = df.sort_values(
        by="composite_quality_score",
        ascending=False
    )

    df = df.reset_index(drop=True)

    return df

def export_results(df, preset_name):

    print()
    print("=" * 60)
    print("EXPORTING RESULTS")
    print("=" * 60)

    output_file = (
        f"output/{preset_name}_results.csv"
    )

    df.to_csv(
        output_file,
        index=False
    )

    print()
    print("Exported Successfully")
    print("File :", output_file)


def generate_summary_report(all_df, filtered_df, preset_name):

    print()
    print("=" * 60)
    print("GENERATING SUMMARY REPORT")
    print("=" * 60)

    report_file = (
        f"output/{preset_name}_summary.txt"
    )

    total_companies = len(all_df)

    filtered_companies = len(filtered_df)

    success_rate = (
        filtered_companies / total_companies
    ) * 100

    avg_score = round(
        filtered_df["composite_quality_score"].mean(),
        2
    )

    with open(report_file, "w") as file:

        file.write("NIFTY100 STOCK SCREENER REPORT\n")
        file.write("=" * 60 + "\n\n")

        file.write(
            f"Generated On : "
            f"{datetime.now()}\n"
        )

        file.write(
            f"Preset : {preset_name}\n\n"
        )

        file.write(
            f"Total Companies : {total_companies}\n"
        )

        file.write(
            f"Filtered Companies : {filtered_companies}\n"
        )

        file.write(
            f"Success Rate : {success_rate:.2f}%\n"
        )

        file.write(
            f"Average Composite Score : {avg_score}\n\n"
        )

        file.write("TOP 10 COMPANIES\n")
        file.write("-" * 60 + "\n")

        top10 = filtered_df.head(10)

        for _, row in top10.iterrows():

            file.write(
                f"{row['company_id']} | "
                f"{row['year']} | "
                f"{row['composite_quality_score']:.2f}\n"
            )

    print()
    print("Summary Report Created")
    print("File :", report_file)


def sector_summary(df):

    print()
    print("=" * 60)
    print("SECTOR SUMMARY")
    print("=" * 60)

    sector_df = (
        df.groupby("broad_sector")
        .agg(
            company_count=("company_id", "count"),
            avg_score=("composite_quality_score", "mean"),
            max_score=("composite_quality_score", "max"),
            min_score=("composite_quality_score", "min")
        )
        .reset_index()
    )

    # Round scores
    sector_df["avg_score"] = sector_df["avg_score"].round(2)
    sector_df["max_score"] = sector_df["max_score"].round(2)
    sector_df["min_score"] = sector_df["min_score"].round(2)

    # Sort by average score
    sector_df = sector_df.sort_values(
        by="avg_score",
        ascending=False
    )

    sector_df = sector_df.reset_index(drop=True)

    print()
    print(sector_df)

    return sector_df


def export_sector_summary(sector_df, preset_name):

    print()
    print("=" * 60)
    print("EXPORTING SECTOR SUMMARY")
    print("=" * 60)

    output_file = f"output/{preset_name}_sector_summary.csv"

    sector_df.to_csv(
        output_file,
        index=False
    )

    print()
    print("Sector Summary Exported Successfully")
    print("File :", output_file)


def top_companies_by_sector(df):

    print()
    print("=" * 60)
    print("TOP COMPANIES BY SECTOR")
    print("=" * 60)

    # Sort by sector and score
    df = df.sort_values(
        by=["broad_sector", "composite_quality_score"],
        ascending=[True, False]
    )

    # Top 3 companies from each sector
    top_companies = (
        df.groupby("broad_sector")
        .head(3)
        .reset_index(drop=True)
    )

    for sector in top_companies["broad_sector"].unique():

        print()
        print("-" * 50)
        print(sector)
        print("-" * 50)

        sector_df = top_companies[
            top_companies["broad_sector"] == sector
        ]

        print(
            sector_df[
                [
                    "company_id",
                    "year",
                    "composite_quality_score"
                ]
            ]
        )

    return top_companies


def export_top_companies(top_companies, preset_name):

    print()
    print("=" * 60)
    print("EXPORTING TOP COMPANIES")
    print("=" * 60)

    output_file = (
        f"output/{preset_name}_top_sector_companies.csv"
    )

    top_companies.to_csv(
        output_file,
        index=False
    )

    print()
    print("Top Companies Exported Successfully")
    print("File :", output_file)


def generate_sector_insights(
    result,
    sector_df,
    preset_name
):

    print()
    print("=" * 60)
    print("GENERATING SECTOR INSIGHTS")
    print("=" * 60)

    report_file = (
        f"output/{preset_name}_sector_insights.txt"
    )

    top_company = result.iloc[0]

    with open(report_file, "w") as file:

        file.write("SECTOR INSIGHTS REPORT\n")
        file.write("=" * 60 + "\n\n")

        file.write(f"Preset : {preset_name}\n")
        file.write(f"Total Filtered Companies : {len(result)}\n")
        file.write(f"Total Sectors : {sector_df.shape[0]}\n\n")

        file.write("Best Performing Sector\n")
        file.write("----------------------\n")
        file.write(
            f"{sector_df.iloc[0]['broad_sector']}\n"
        )

        file.write(
            f"Average Score : "
            f"{sector_df.iloc[0]['avg_score']}\n\n"
        )

        file.write("Top Company\n")
        file.write("----------------------\n")
        file.write(
            f"Company : {top_company['company_id']}\n"
        )
        file.write(
            f"Year : {top_company['year']}\n"
        )
        file.write(
            f"Score : "
            f"{round(top_company['composite_quality_score'],2)}\n"
        )

    print()
    print("Sector Insights Report Generated")
    print("File :", report_file)


def choose_sort_column():

    print()
    print("=" * 60)
    print("SORT OPTIONS")
    print("=" * 60)

    options = {
        "1": "composite_quality_score",
        "2": "return_on_equity_pct",
        "3": "revenue_cagr_5yr",
        "4": "pat_cagr_5yr",
        "5": "operating_profit_margin_pct"
    }

    print("1. Composite Quality Score")
    print("2. Return on Equity")
    print("3. Revenue CAGR (5Y)")
    print("4. PAT CAGR (5Y)")
    print("5. Operating Profit Margin")

    choice = input("\nChoose sorting option: ").strip()

    return options.get(choice, "composite_quality_score")


def choose_top_n():

    print()
    print("=" * 60)
    print("TOP N RESULTS")
    print("=" * 60)

    number = input("Enter Number of Companies : ")

    try:
        return int(number)

    except:

        return 20


def export_all_presets(config):

    print()
    print("=" * 60)
    print("GENERATING SCREENER_OUTPUT.XLSX")
    print("=" * 60)

    with pd.ExcelWriter(
        "output/screener_output.xlsx",
        engine="openpyxl"
    ) as writer:

        for preset_name, filters in config.items():

            print("Processing :", preset_name)

            df = load_financial_ratios()

            result = apply_filters(
                df,
                {"filters": filters}
            )

            result = result.sort_values(
                "composite_quality_score",
                ascending=False
            )

            result.to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

    print()
    print("Workbook Created Successfully")
    print("File : output/screener_output.xlsx")


def apply_conditional_formatting(config):

    print()
    print("=" * 60)
    print("APPLYING CONDITIONAL FORMATTING")
    print("=" * 60)

    workbook = load_workbook("output/screener_output.xlsx")

    for preset_name, filters in config.items():

        sheet = workbook[preset_name]

        headers = {}

        for cell in sheet[1]:
            headers[cell.value] = cell.column

        for row in range(2, sheet.max_row + 1):

            # ROE
            if "roe_min" in filters:

                col = headers.get("return_on_equity_pct")

                if col:

                    cell = sheet.cell(row=row, column=col)

                    if cell.value >= filters["roe_min"]:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

            # Debt
            if "debt_to_equity_max" in filters:

                col = headers.get("debt_to_equity")

                if col:

                    cell = sheet.cell(row=row, column=col)

                    if cell.value <= filters["debt_to_equity_max"]:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

            # Revenue CAGR
            if "revenue_cagr_5yr_min" in filters:

                col = headers.get("revenue_cagr_5yr")

                if col:

                    cell = sheet.cell(row=row, column=col)

                    if cell.value >= filters["revenue_cagr_5yr_min"]:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

            # PAT CAGR
            if "pat_cagr_5yr_min" in filters:

                col = headers.get("pat_cagr_5yr")

                if col:

                    cell = sheet.cell(row=row, column=col)

                    if cell.value >= filters["pat_cagr_5yr_min"]:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

            # Free Cash Flow
            if "free_cash_flow_min" in filters:

                col = headers.get("free_cash_flow_cr")

                if col:

                    cell = sheet.cell(row=row, column=col)

                    if cell.value >= filters["free_cash_flow_min"]:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

    workbook.save("output/screener_output.xlsx")

    print()
    print("Conditional Formatting Applied Successfully")

if __name__ == "__main__":

    config = load_config()

    print()

    print("Available Presets")

    for preset in config.keys():
      print("-", preset)

    print()

    selected = input("Enter Preset : ").strip().lower()

    filters = config[selected]


    df = load_financial_ratios()

    print()
    print("Columns Available")
    print(df.columns.tolist())

    result = apply_filters(df, {"filters": filters})
    sort_column = choose_sort_column()
    top_n = choose_top_n()

    result = result.sort_values(
    by=sort_column,
    ascending=False
    ).reset_index(drop=True)

    print()

    print("Rows :", len(result))

    print()

    print("=" * 60)
    print("TOP FILTERED COMPANIES")
    print("=" * 60)

    print(result.head(top_n))

    export_results(
    result.head(top_n),
    selected
    )

    generate_summary_report(
    df,
    result.head(top_n),
    selected
    )

    sector_df = sector_summary(result)

    export_sector_summary(
    sector_df,
    selected
    )

    top_sector_companies = top_companies_by_sector(result)

    export_top_companies(
    top_sector_companies,
    selected
    )

    generate_sector_insights(
    result,
    sector_df,
    selected
    )

    export_all_presets(config)

    apply_conditional_formatting(config)